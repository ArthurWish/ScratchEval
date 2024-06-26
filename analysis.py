import json
from antlr4 import *
from ScratchLexer import ScratchLexer
from ScratchParser import ScratchParser
from ScratchListener import ScratchListener
import zipfile
import os
import pickle
import codecs
# workbook相关
from openpyxl.workbook import Workbook
# 一个eggache的数字转为列字母的方法
from openpyxl.utils import get_column_letter
import time
from collections import Counter
from pathlib import Path
import networkx as nx
import matplotlib.pyplot as plt
from networkx.readwrite import json_graph

class HandleExcel():
    def __init__(self):
        self.head_row_labels = [u'Name', u'Abstraction', u'Parallelism', u'LogicalThinking', u'Synchronization',
                                u'FlowControl', u'UserInteractivity', u'DataRepresentation', u'CodeOrganization']
        # self.head_row_labels = [u'Name', u'Abstraction', u'Parallelism', u'LogicalThinking', u'Synchronization',
        #                         u'FlowControl', u'UserInteractivity', u'DataRepresentation', u'CodeOrganization',
        #                         u'n1', u'n2', u'N1', u'N2', u'CC']
    def read_from_file(self):
        info = {}
        path = os.path.abspath('.')
        filepath = path + '/test'
        filepath = path + '/graph'
        filepath = path + '/error'
        # filepath = "/test"
        # filepath = "/12.22思维导图测试案例—12人"
        pathDir = Path(filepath).glob("*.sb3")
        # pathDir = ["/media/sda1/cyn-workspace/Scratch-project/王致远Scratch作品.sb3", "/media/sda1/cyn-workspace/Scratch-project/王致远ChatScratch作品.sb3"]
        
        for allDir in pathDir:
            child = os.path.join(filepath, allDir)
            print("name=", allDir)
            listener_info = ctAnalysis(child)
            print("listener_info", listener_info)
            # listener_operators = gen(child)[3]
            # # print("listener_operators", listener_operators)
            # listener_operand = gen(child)[4]
            # # print("listener_operand", listener_operand)
            # listener_cc = gen(child)[5]
            # # print("listener_CC", listener_cc)
            #
            # n1 = len(set(listener_operators))
            # # print("set listener_operators", set(listener_operators))
            # N1 = len(listener_operators)
            # n2 = len(set(listener_operand))
            # # print("set listener_operand", set(listener_operand))
            # N2 = str(len(listener_operand))
            #
            # listener_info['n1'] = n1
            # listener_info['n2'] = n2
            # listener_info['N1'] = N1
            # listener_info['N2'] = N2
            #
            # listener_info['cc'] = listener_cc

            info[allDir] = listener_info
            # print("info[allDir]", info[allDir])

        return info

    def write_to_excel_with_openpyxl(self, records, head_row, save_excel_name):
        # 新建一个workbook
        wb = Workbook()
        
        # 设置文件输出路径与名称
        dest_filename = save_excel_name
        
        # 第一个sheet是ws
        ws = wb.active
        ws.title = "range names"
        
        # 写第一行，标题行
        for h_x in range(1, len(head_row) + 1):
            h_col = get_column_letter(h_x)
            ws['%s%s' % (h_col, 1)].value = head_row[h_x - 1]
        
        # 写第二行及其以后的那些行
        row = 2
        for name, points in records.items():
            ws['A%d' % row].value = str(name)
            col = 2
            for point in points:
                col_letter = get_column_letter(col)
                ws['%s%d' % (col_letter, row)].value = points[point]
                col += 1
            row += 1
        
        wb.save(filename=dest_filename)


    def run_main_save_to_excel_with_openpyxl(self):
        dataset_list = self.read_from_file()
        '''test use openpyxl to handle EXCEL 2007'''
        head_row_label = self.head_row_labels
        save_name = "test_openpyxl.xlsx"
        self.write_to_excel_with_openpyxl(dataset_list, head_row_label, save_name)
def unzip_scratch(filename):
    """
    unzip scratch project and extract project.json file
    :param filename: filename fo scratch project
    :return: null or project.json content
    """
    zfile = zipfile.ZipFile(filename, 'r')
    if "project.json" in zfile.namelist():
        data = zfile.read("project.json")
        json_data = json.loads(data)
        with open('./test.json', 'w', encoding='utf-8') as json_file:
            json.dump(json_data, json_file, ensure_ascii=False, indent=4)
        return data
    else:
        return None

def ctAnalysis(argv):
    raw_json = unzip_scratch(argv)
    encoded_json = codecs.decode(raw_json, 'utf-8', 'strict')
    input = InputStream(encoded_json)
    lexer = ScratchLexer(input)
    stream = CommonTokenStream(lexer)
    parser = ScratchParser(stream)
    tree = parser.json()
    walker = ParseTreeWalker()
    listener = ScratchListener()
    walker.walk(listener, tree) # 结束该函数后，得到所有以 dict形式的 block信息
    ct_score = {}
    ct_score['abstract'] = 0
    ct_score['parallel'] = 0
    ct_score['logic'] = 0
    ct_score['synchron'] = 0
    ct_score['flowControl'] = 0   
    ct_score['userInteraction'] = 0
    ct_score['dataRepresent'] = 1
    ct_score['codeOrganize'] = 0
    block_count = listener.block_count  # block数目
    sprite_count = listener.sprite_count # sprite数目
    comment_count = listener.comment_count  # 注释数目
    block = listener.id_block  # block
    # block_json = json.loads(block)
    # with open('./test2.json', 'w', encoding='utf-8') as json_file:
    #         json.dump(block, json_file, ensure_ascii=False, indent=4)
    broadcast_received = listener.broadcast_received
    broadcast = listener.broadcast
    listdic = listener.list
    # print('listdic',listdic)
    listlen = {}
    for i in listdic.keys():
        # print('i!!!!', i)
        listlen[i] = len(listdic[i])
    # print('listlen!!!!',listlen)
    isQueue = listdic.copy()
    # print('isQueue!!!!',isQueue)
    isStack = listdic.copy()
    deadBlock_count = 0
    flagClicked_count = 0
    keyPressed_count = 0
    spriteClicked_count = 0
    backdropSwitch_count = 0
    backClicked_count = 0
    greaterThan_count = 0
    broadcastRecive_dicc = Counter()
    motionGreater_count = 0
    logic_operation = {'operator_and', 'operator_or', 'operator_not'}
    sprite_init = 1
    changeset=set()
    between_block_dict={}
    overp=set()
    changedict={}
    
    with open('changeset.pkl', 'rb') as file:
        change_set = pickle.load(file)
    with open('blockdict.json', 'rb') as file:
        block_dict = json.load(file)
        
    # num=0    
    # for atr in change_set:
    #     print(str(num)+" "+atr)
    #     num=num+1
    #print("changedict")
    #print(changedict)
    
    block_grade={}
    with open('blockgrade.json', 'rb') as file:
        block_grade = json.load(file)
        
    block_point={}
    
    # num=0 
    # for d in block_dict.keys():
    #     print(str(num)+" "+d)
    #     num=num+1
    # print("block_grade")
    #print(block_grade) 
    # 遍历所有block，给出ct得分
    
    #克隆等集合(start为启动块) 
    
    
    start_clone_set=set()
    start_broad_set=set()
    start_backdrop_set=set()
    start_func_set=set()
    start_sound_set=set()
    num=1
    graphid={}
    
    
    for id in block.keys():
        graphid[id]=num
        num=num+1
        isDead_code = block[id].getIsDead()
        name = block[id].getName()
        sprite=block[id].getSpriteName()
        parent_id = block[id].getParent()
        if(parent_id==None and block[id].name in block_dict.values()):
            if(block[id].name.find("clone")!=-1):
                start_clone_set.add(id)
            else :
                if(block[id].name.find("broad")!=-1):
                        start_broad_set.add(id)
                else :
                    if(block[id].name.find("backdrop")!=-1):
                        start_backdrop_set.add(id)
                    else :
                        if(block[id].name.find("call")!=-1):
                            start_func_set.add(id)
                        else:
                            start_sound_set.add(id)
                    

        if(parent_id==None and id not in overp):
            overp.add(id)
            point=1
            if(block[id].name in block_dict.values()):
                point=block_grade[list(block_dict.keys())[list(block_dict.values()).index(block[id].name)]]  

            def set_point(id,point):
                if(id==None):return
                block_point[id]=point
                #print(id+" "+str(point))
                set_point(block[id].getNext(),point)
                set_point(block[id].getSubstack(),point)
                set_point(block[id].getSubstack2(),point)
                
            set_point(id,point)

        
                #print(block[id].name+" "+str(point))
        # if(sprite=="角色1"):
        #     changeset.add(name)
        # if(sprite=="角色2"):
        #     if(parent_id==None and id not in overp):
        #         overp.add(id)
        #         nowblock=block[id].getNext()
        #         while(nowblock!=None):
        #             between_block_dict[block[nowblock].name]=block[id].name
        #             nowblock=block[nowblock].getNext()
        
        # 父block是deadcode则子block一定是deadcode
        # note: block 下面有substack也是属于block，需要判断
        # if parent_id is not None and parent_id in block.keys() and block[parent_id].getIsDead() is True:
        #     block[id].setIsDead(True)
        #     deadBlock_count += 1
        #     substack = block[id].getSubstack()
        #     if substack is not None and substack in block.keys():
        #         block[substack].setIsDead(True)
        #         deadBlock_count += 1
        #     substack2 = block[id].getSubstack2()
        #     if substack2 is not None and substack2 in block.keys():
        #         block[substack2].setIsDead(True)
        #         deadBlock_count += 1
        #     next_id = block[id].getNext()
            
            # while next_id is not None and next_id in block.keys(): # 如果该模块是dead，那么下面的所有模块都设置为dead
            #     block[next_id].setIsDead(True)
            #     deadBlock_count += 1
            #     substack = block[next_id].getSubstack()
            #     if substack is not None and substack in block.keys():
            #         block[substack].setIsDead(True)
            #         deadBlock_count += 1
            #     substack2 = block[id].getSubstack2()
            #     if substack2 is not None and substack2 in block.keys():
            #         block[substack2].setIsDead(True)
            #         deadBlock_count += 1
            #     next_id = block[next_id].getNext()
             
        # 子block继承父block的proc
        if parent_id is not None and parent_id in block.keys() and block[parent_id].getProcId() is not None and block[id].getProcId() is not None:
            proc_id = block[parent_id].getProcId()
            block[id].setProcId(proc_id)
            substack = block[id].getSubstack()
            if substack is not None and substack in block.keys():
                block[substack].setProcId(proc_id)
            substack2 = block[id].getSubstack2()
            if substack2 is not None and substack2 in block.keys():
                block[substack2].setProcId(proc_id)
            next_id = block[id].getNext()
            while next_id is not None and next_id in block.keys():
                block[next_id].setProcId(proc_id)
                substack = block[next_id].getSubstack()
                if substack is not None and substack in block.keys():
                    block[substack].setProcId(proc_id)
                substack2 = block[id].getSubstack2()
                if substack2 is not None and substack2 in block.keys():
                    block[substack2].setProcId(proc_id)
                next_id = block[next_id].getNext()
        # 判断起始语句是否正确，不正确则为deadcode
        if isDead_code is False:
            if parent_id is None:  # 判断block的起始语句是否为启动语句
                if name.find("when") == -1 and name.find("start") == -1 and name.find("definition") == -1:
                    # block[id].setIsDead(True) # 如果不是启动语句则整个script是dead
                    # deadBlock_count += 1
                    # substack = block[id].getSubstack()
                    # if substack is not None and substack in block.keys():
                    #     block[substack].setIsDead(True)
                    #     deadBlock_count += 1
                    # substack2 = block[id].getSubstack2()
                    # if substack2 is not None and substack2 in block.keys():
                    #     block[substack2].setIsDead(True)
                    #     deadBlock_count += 1
                    # 将所有的next block设置为deadcode
                    deadBlock_count
                    def setdead(id):
                        if(id==None):return
                        block[id].setIsDead(True)
                        nonlocal deadBlock_count
                        deadBlock_count+=1
                        #print(id+" "+str(point))
                        setdead(block[id].getNext())
                        setdead(block[id].getSubstack())
                        setdead(block[id].getSubstack2())
                        
                    setdead(id)  
                    
                    # next_id = block[id].getNext()
                    # while next_id is not None and next_id in block.keys():
                    #     block[next_id].setIsDead(True)
                    #     deadBlock_count += 1
                    #     substack = block[next_id].getSubstack()
                    #     if substack is not None and substack in block.keys():
                    #         block[substack].setIsDead(True)
                    #         deadBlock_count += 1
                    #     substack2 = block[next_id].getSubstack2()
                    #     if substack2 is not None and substack2 in block.keys():
                    #         block[substack2].setIsDead(True)
                    #         deadBlock_count += 1
                    #     next_id = block[next_id].getNext()
        # proc记录
        if name == 'procedures_definition':
            block[id].setProcId(id)
            substack = block[id].getSubstack()
            if substack is not None and substack in block.keys():
                block[substack].setProcId(id)
            substack2 = block[id].getSubstack2()
            if substack2 is not None and substack2 in block.keys():
                block[substack2].setProcId(id)
            next_id = block[id].getNext()
            while next_id is not None and next_id in block.keys():
                block[next_id].setProcId(id)
                substack = block[next_id].getSubstack()
                if substack is not None and substack in block.keys():
                    block[substack].setProcId(id)
                substack2 = block[id].getSubstack2()
                if substack2 is not None and substack2 in block.keys():
                    block[substack2].setProcId(id)
                next_id = block[next_id].getNext()
        if isDead_code is False:
            if name == 'event_whengreaterthan':
                greaterThan_count += 1
            if name == 'videoSensing_whenMotionGreaterThan':
                motionGreater_count += 1
            if name == 'event_whenbroadcastreceived' and block[id].getBroadcast() in broadcast:
                # 评分标准 4-5
                if ct_score['synchron'] < 5:
                    ct_score['synchron'] = 5
                broadcastRecive_dicc[block[id].getBroadcast()] += 1

            if name == 'event_whenkeypressed':
                # 评分标准4-2
                if ct_score['synchron'] < 2:
                    ct_score['synchron'] = 2
                # 评分标准6-3
                if ct_score['userInteraction'] < 3:
                    ct_score['userInteraction'] = 3
                keyPressed_count += 1
            if name == 'event_whenthisspriteclicked':
                # 评分标准4-2
                if ct_score['synchron'] < 2:
                    ct_score['synchron'] = 2
                # 评分标准6-3
                if ct_score['userInteraction'] < 3:
                    ct_score['userInteraction'] = 3
                spriteClicked_count += 1
            if name == 'event_whenbackdropswitchesto':
                backdropSwitch_count += 1
                # 评分标准4-2
                if ct_score['synchron'] < 2:
                    ct_score['synchron'] = 2
            if name == 'event_whenstageclicked':
                # 评分标准4-2
                if ct_score['synchron'] < 2:
                    ct_score['synchron'] = 2
                # 评分标准6-3
                if ct_score['userInteraction'] < 3:
                    ct_score['userInteraction'] = 3
                backClicked_count += 1
            if name == 'event_whenflagclicked':
                # 评分标准6-2
                if ct_score['userInteraction'] < 2:
                    ct_score['userInteraction'] = 2
                flagClicked_count += 1

            # 列表操作
            if name == 'data_deleteoflist':
                loc = block[id].getLocation()
                if type(loc) is int:
                    listname = block[id].getListName()
                    if (listname in isQueue.keys()) and loc != 1:
                        del isQueue[listname]

                    if (listname in isStack.keys()) and loc != listlen[listname]:
                        del isStack[listname]

                    if listname in listdic.keys() and listlen[listname] >= loc:
                        listdic[listname].pop(loc - 1)
                        # print("删除list", listname, loc, listdic[listname ])
                        listlen[listname] -= 1
            if name == 'data_addtolist':
                # print('list', listdic.items())
                listname = block[id].getListName()
                # print('listname', listname)
                content = block[id].getListContent()
                if listname in listdic.keys():
                    listdic[listname].append(content)
                    listlen[listname] += 1
                    # print("添加list", listname, listdic[listname])
            if name == 'data_insertatlist':
                loc = block[id].getLocation()
                if type(loc) is int:
                    listname = block[id].getListName()
                    content = block[id].getListContent()
                    if (listname in isQueue.keys()) and loc != listlen[listname] + 1:
                        del isQueue[listname]

                    if (listname in isStack.keys()) and loc != listlen[listname] + 1:
                        del isStack[listname]

                    if listname in listdic.keys() and loc <= listlen[listname]:
                        listdic[listname].insert(loc - 1, content)
                        # print("插入list", listname, loc, listdic[listname])
                        listlen[listname] += 1
            if name == 'data_replaceitemoflist':
                loc = block[id].getLocation()

                if type(loc) is int:
                    listname = block[id].getListName()
                    # print("替换list！！！", listname, loc, listlen[listname], listdic[listname])
                    content = block[id].getListContent()
                    if listname in isQueue.keys():
                        del isQueue[listname]

                    if listname in isStack.keys():
                        del isStack[listname]

                    if listname in listdic.keys() and loc <= listlen[listname]:
                        listdic[listname][loc - 1] = content
                        # print("替换list", listname, loc, listdic[listname])

            # 角色属性初始化
            if name == 'looks_nextcostume':
                isfind = 0
                i = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():

                    if block[parent_id].getName() == 'looks_switchcostumeto':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                    i += 1
                if isfind == 0:
                    sprite_init = 0
            if name == 'looks_changesizeby':
                isfind = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():
                    if block[parent_id].getName() == 'looks_setsizeto':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                if isfind == 0:
                    sprite_init = 0
            if name == 'looks_nextbackdrop':
                isfind = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():
                    if block[parent_id].getName() == 'looks_switchbackdropto':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                if isfind == 0:
                    sprite_init = 0
            if name in {'motion_turnright', 'motion_turnleft', 'motion_pointtowards'}:
                isfind = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():
                    if block[parent_id].getName() == 'motion_pointindirection':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                if isfind == 0:
                    sprite_init = 0
            if name in {'motion_movesteps', 'motion_glideto', 'motion_glidesecstoxy', 'motion_goto'}:
                isfind = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():
                    if block[parent_id].getName() == 'motion_gotoxy':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                if isfind == 0:
                    sprite_init = 0

            # 评分标准1-5
            if ct_score['abstract'] < 5 and name == 'procedures_call' and block[id].getProcId() is not None:
                ct_score['abstract'] = 5
            # 评分标准1-4
            if ct_score['abstract'] < 4 and name.find('clone') != -1:
                ct_score['abstract'] = 4
            # 评分标准1-3
            if ct_score['abstract'] < 3 and name == 'procedures_call':
                ct_score['abstract'] = 3
            # 评分标准1-2
            if ct_score['abstract'] < 2 and name in{'looks_switchbackdroptoandwait', 'looks_backdrops',
                                                    'looks_nextbackdrop', 'looks_nextcostume', 'looks_switchcostumeto'} :
                ct_score['abstract'] = 2

            if name == 'control_if' or name == 'control_if_else':
                substack = block[id].getSubstack()
                isfind = 0
                while isfind == 0 and substack is not None and substack in block.keys():
                    if block[substack].getName() in {'control_if', 'control_if_else'}:
                        isfind = 1
                    if block[substack].getName() in {'control_repeat', 'control_forever', 'control_repeat_until'}:
                        isfind = 2
                    substack = block[substack].getNext()
                # 评分标准3-5
                if ct_score['logic'] < 5 and isfind == 2:
                    ct_score['logic'] = 5
                # 评分标准3-4
                if ct_score['logic'] < 4 and isfind == 1:
                    ct_score['logic'] = 4
                # 评分标准3-3
                if ct_score['logic'] < 3:
                    condition = block[id].getCondition()
                    if condition in logic_operation:
                        ct_score['logic'] = 3
            if name == 'control_if_else':
                substack2 = block[id].getSubstack2()
                isfind = 0
                while isfind == 0 and substack2 is not None and substack2 in block.keys():
                    if block[substack2].getName() in {'control_if', 'control_if_else'}:
                        isfind = 1
                    if block[substack2].getName() in {'control_repeat', 'control_forever', 'control_repeat_until'}:
                        isfind = 2
                    substack2 = block[substack2].getNext()
                # 评分标准3-5
                if ct_score['logic'] < 5 and isfind == 2:
                    ct_score['logic'] = 5
                # 评分标准3-4
                if ct_score['logic'] < 4 and isfind == 1:
                    ct_score['logic'] = 4

            # 评分标准3-1
            if ct_score['logic'] == 0 and name == 'control_if':
                ct_score['logic'] = 1
            # 评分标准3-2
            if ct_score['logic'] < 2 and name == 'control_if_else':
                ct_score['logic'] = 2

            # 评分标准4-4
            if ct_score['synchron'] < 4 and name == 'control_wait_until':
                ct_score['synchron'] = 4
            # 评分标准4-3
            if ct_score['synchron'] < 3 and name in {'sensing_touchingcolor', 'sensing_coloristouchingcolor',
                                                     'sensing_loudness', 'sensing_timer', 'sensing_current',
                                                     'videoSensing_videoOn', 'sensing_dayssince2000', 'sensing_of'}:
                ct_score['synchron'] = 3
            # 评分标准4-2
            if ct_score['synchron'] < 2 and name in {'sensing_touchingobject', 'sensing_keypressed', 'sensing_mousedown',
                                                     'sensing_mousex', 'sensing_mousey', 'sensing_distanceto', 'looks_costumenumbername'}:
                ct_score['synchron'] = 2
            # 评分标准4-1
            if ct_score['synchron'] == 0 and name in {'control_wait', 'control_stop'}:
                ct_score['synchron'] = 1

            if name in {'control_repeat', 'control_forever', 'control_repeat_until'}:
                substack = block[id].getSubstack()
                isfind = 0
                while isfind == 0 and substack is not None and substack in block.keys():
                    if block[substack].getName() in {'control_if', 'control_if_else', 'control_repeat',
                                                     'control_forever', 'control_repeat_until'}:
                        isfind = 1
                    substack = block[substack].getNext()
                # 评分标准5-5
                if ct_score['flowControl'] < 5 and isfind == 1:
                    ct_score['flowControl'] = 5
                # 评分标准5-4
                if ct_score['flowControl'] < 4:
                    condition = block[id].getCondition()
                    if condition in logic_operation:
                        ct_score['flowControl'] = 4
            if name == 'control_repeat_until':
                # 评分标准5-3
                if ct_score['flowControl'] < 3:
                    ct_score['flowControl'] = 3
            # 评分标准5-2
            if ct_score['flowControl'] < 2 and name in {'control_repeat', 'control_forever'}:
                ct_score['flowControl'] = 2

            # 评分标准6-5
            if ct_score['userInteraction'] < 5 and name == 'sensing_askandwait':
                ct_score['userInteraction'] = 5
                # 评分标准6-4
            if ct_score['userInteraction'] < 4 and name in {'videoSensing_whenMotionGreaterThan',
                                                            'videoSensing_videoOn', 'sensing_loudness'}:
                ct_score['userInteraction'] = 4
            # 评分标准6-3
            if ct_score['userInteraction'] < 3 and name in {'sensing_keypressed', 'sensing_mousedown'}:
                ct_score['userInteraction'] = 3
            # 评分标准6-1
            if ct_score['userInteraction'] == 0 and name in {'looks_sayforsecs', 'looks_say', 'looks_thinkforsecs', 'looks_think'}:
                ct_score['userInteraction'] = 1

            # 评分标准7-4
            if ct_score['dataRepresent'] < 4 and name in {'data_addtolist', 'data_deleteoflist',
                                                          'data_deletealloflist', 'data_insertatlist',
                                                          'data_replaceitemoflist', 'data_itemoflist',
                                                          'data_itemnumoflist', 'data_lengthoflist',
                                                          'data_listcontainsitem', 'data_showlist',
                                                          'data_hidelist'}:
                ct_score['dataRepresent'] = 4
            # 评分标准7-3
            if ct_score['dataRepresent'] < 3 and name in {'data_setvariableto', 'data_changevariableby',
                                                          'data_showvariable', 'data_hidevariable'}:
                ct_score['dataRepresent'] = 3
            # 评分标准7-2
            if ct_score['dataRepresent'] < 2 and name == 'operator_join':
                ct_score['dataRepresent'] = 2


    print(len(block))
    
    
    g=nx.Graph()
    dg=nx.DiGraph()
    oversprite=set()
    save_graph=[]
    for id in block.keys():
        idn=graphid[id]
        #print(idn)
        if(block[id].isDead==True) :continue
        if(id not in block_point.keys()):
            continue
            
        point=block_point[id]
        
        
        if(block[id].spriteName not in oversprite):
            g.add_weighted_edges_from([(block[id].spriteName,"start_flag",1)])
            oversprite.add(block[id].spriteName)
            save_graph.append({"source":block[id].spriteName,"target":"start_flag","weight":1})
          
        name=block[id].name
        
        if not (name.find("when") == -1 and name.find("start") == -1 ):
            g.add_weighted_edges_from([(block[id].spriteName,idn,point)])
            save_graph.append({"source":block[id].spriteName,"target":idn,"weight":point})
        if not name.find("definition")== -1:
            #print(idn)
            g.add_weighted_edges_from([("start_flag",idn,point)])
            save_graph.append({"source":"start_flag","target":idn,"weight":point})
        #和next和substack建立连接
        if(block[id].getNext()!=None):
            g.add_weighted_edges_from([(graphid[block[id].getNext()],idn,point)])
            save_graph.append({"source":graphid[block[id].getNext()],"target":idn,"weight":point})
        if(block[id].getSubstack()!=None):
            g.add_weighted_edges_from([(graphid[block[id].getSubstack()],idn,point)])
            save_graph.append({"source":graphid[block[id].getSubstack()],"target":idn,"weight":point})
        if(block[id].getSubstack2()!=None):
            g.add_weighted_edges_from([(graphid[block[id].getSubstack2()],idn,point)])
            save_graph.append({"source":graphid[block[id].getSubstack2()],"target":idn,"weight":point})
        #block之间建立连接
        if(name in change_set):
            g.add_weighted_edges_from([(block[id].spriteName,idn,point)])
            save_graph.append({"source":block[id].spriteName,"target":idn,"weight":point})
        if(name in block_dict.keys()):
             if(name.find("clone")!=-1):
                 for id2 in start_clone_set:
                     if(block[id2].getSpriteName()==block[id].getSpriteName()):
                        g.add_weighted_edges_from([(idn,graphid[id2],block_grade[name])])
                        save_graph.append({"source":graphid[id2],"target":idn,"weight":block_grade[name]})
             if(name.find("broadcast")!=-1):
                 for id2 in start_broad_set:
                     if(block[id2].broadcast==block[id].broadcast):
                        g.add_weighted_edges_from([(idn,graphid[id2],block_grade[name])])
                        save_graph.append({"source":graphid[id2],"target":idn,"weight":block_grade[name]})
             if(name.find("backdrop")!=-1):
                  for id2 in start_backdrop_set:
                        g.add_weighted_edges_from([(idn,graphid[id2],block_grade[name])])
                        save_graph.append({"source":graphid[id2],"target":idn,"weight":block_grade[name]})
             if(name.find("call")!=-1):
                 for id2 in start_func_set:
                     if(block[id2].proc_id==block[id].proc_id):
                        g.add_weighted_edges_from([(idn,graphid[id2],block_grade[name])])
                        save_graph.append({"source":graphid[id2],"target":idn,"weight":block_grade[name]})
             if(name.find("sound")!=-1):
                  for id2 in start_sound_set:
                        g.add_weighted_edges_from([(idn,graphid[id2],block_grade[name])])
                        save_graph.append({"source":graphid[id2],"target":idn,"weight":block_grade[name]})
    # for id1 in block.keys():
    #      for id2 in block.keys():
             
           
    #         if(block[id2].isDead==True or block[id1].isDead==True):
    #            continue
    #         g.add_edge(block[id2].spriteName,"start_flag")
    #         name=block[id2].name
    #         if not (name.find("when") == -1 and name.find("start") == -1 ):
    #             g.add_edge(block[id2].spriteName,id2)
                
    #         if not name.find("definition")== -1:
    #             g.add_edge("start_flag",id2)
                
                
    #         if(block[id2].name=="event_whenflagcliked"):
    #             g.add_edge(id2,"start_flag")
    #             dg.add_edge(id2,"start_flag")
    #             continue
    #          #顺序建边
    #         if(id2==block[id1].next ):
    #             g.add_edge(id1,id2)
    #             dg.add_edge(id1,id2)
    #         if(block[id1].substack!=None):
    #             g.add_edge(id1,block[id1].substack)
    #         if(block[id1].substack2!=None):
    #             g.add_edge(id1,block[id1].substack2)
    #         #影响角色建边
    #         if(block[id2].name in change_set):
    #             g.add_edge(id2,block[id2].spriteName)
    #             dg.add_edge(id2,block[id2].spriteName)
    #         #逻辑建边
    #         if(block[id1].name in block_dict.keys()):
    #             if(block[id2].name in block_dict[block[id1].name]):
    #                 if(block[id1].name in {"event_broadcast","event_broadcastandwait"} ):
    #                     if(block[id1].broadcast==block[id2].broadcast):
    #                         g.add_edge(id1,id2)
    #                         dg.add_edge(id1,id2) 
    #                 else:
    #                     if(block[id1].name in {"control_create_clone","control_delete_this_clone"} ):
    #                          if(block[id1].spriteName==block[id2].spriteName):
    #                             g.add_edge(id1,id2)
    #                             dg.add_edge(id1,id2) 
    #                     else:
    #                         g.add_edge(id1,id2)
    #                         dg.add_edge(id1,id2)    
                            
                            
    # with open('changeset.pkl', 'wb') as file:
    #       pickle.dump(changeset, file)
    # with open('changeset.pkl', 'rb') as file:
    #      change_set = pickle.load(file)
    
    
    # with open("blockdict.json", "w", encoding='utf-8') as f:
    # # json.dump(dict_, f)  # 写为一行
    #     json.dump(between_block_dict, f, indent=2, sort_keys=True, ensure_ascii=False)  # 写为多行
    # with open('blockdict.json', 'rb') as file:
    #      block_dict = json.load(file)
    
    
    # 评分标准1-1
    # nx.draw_networkx(g,with_labels=True)
    # plt.show()
    print("undirect")
    print(nx.average_clustering(g))#整个图的聚集系数
    print("---------------------------------------")
    print(nx.average_shortest_path_length(g))#图的平均路径长度
    print("---------------------------------------")
    print(nx.community.modularity(g, nx.community.label_propagation_communities(g)))
    print("---------------------------------------")
    
    # jsg=json.dumps(json_graph.node_link_data(g))
    # f2 = open(argv+'_gragh.json', 'w')
    # f2.write(jsg)
    # f2.close()
    
    # print("有向图")
    # print(nx.average_clustering(dg))#整个图的聚集系数
    # print("---------------------------------------")
    # print(nx.average_shortest_path_length(dg))#图的平均路径长度
    # print("---------------------------------------")
    # print(nx.community.modularity(dg, nx.community.label_propagation_communities(dg)))
    # print("---------------------------------------")
    # print(nx.smallworld(dg))
    
    
    if ct_score['abstract'] == 0 and sprite_count > 1 and block_count > 1:
        ct_score['abstract'] = 1
    # 评分标准5-1
    if ct_score['flowControl'] == 0 and block_count > 1:
        ct_score['flowControl'] = 1
    # 评分标准2-5
    for key in broadcastRecive_dicc.keys():
        if broadcastRecive_dicc[key] > 1 and ct_score['parallel'] < 5:
            ct_score['parallel'] = 5
            break
    # 评分标准2-4
    if ct_score['parallel'] < 4 and (greaterThan_count > 1 or motionGreater_count > 1):
        ct_score['parallel'] = 4
    # 评分标准2-3
    if ct_score['parallel'] < 3 and backdropSwitch_count > 1:
        ct_score['parallel'] = 3
    # 评分标准2-2
    if ct_score['parallel'] < 2 and (keyPressed_count > 1 or spriteClicked_count > 1 or backClicked_count > 1):
        ct_score['parallel'] = 2
    # 评分标准2-1
    if ct_score['parallel'] == 0 and flagClicked_count > 1:
        ct_score['parallel'] = 1

    # 评分标准7-5
    if ct_score['dataRepresent'] < 5 and (len(isQueue) != 0 or len(isStack) != 0):
        ct_score['dataRepresent'] = 5

    if broadcast_received and broadcast:
        isfind = 1
        for message in broadcast:
            if message not in broadcast_received:
                isfind = 0
                break
        # 评分标准8-5
        if ct_score['codeOrganize'] < 5 and isfind == 1:
            ct_score['codeOrganize'] = 5
    # 评分标准8-4
    if ct_score['codeOrganize'] < 4 and deadBlock_count == 0:
        ct_score['codeOrganize'] = 4
    # 评分标准8-3
    if ct_score['codeOrganize'] < 3 and comment_count > 0:
        ct_score['codeOrganize'] = 3
    # 评分标准8-1
    if ct_score['codeOrganize'] == 0 and sprite_init == 1:
        ct_score['codeOrganize'] = 1

    return ct_score



if __name__ == '__main__':
    start_time = time.time()
   
    p = HandleExcel()
    p.run_main_save_to_excel_with_openpyxl()
    # ctAnalysis(sys.argv[1])
    end_time = time.time()
    spend_time = end_time - start_time
    print("time:%.4f" % spend_time)
